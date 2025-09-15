# dienstplan_empfehlung.py
# Streamlit-App: Dienstpläne (Blatt "Touren") einlesen, analysieren und
# Excel-Output mit 2 Blättern erzeugen:
# 1) Einsatz Folgewoche (So–Sa): Name + Startzeit-Empfehlungen (HH:MM)
# 2) Typische Startzeiten: globale/wochentags-Mediane, Sample-Zahlen, Rotation-Hinweis

import io
from datetime import datetime, date, time, timedelta
from typing import List, Optional, Tuple, Dict

import numpy as np
import pandas as pd
import streamlit as st

# =========================
# ------- Helpers ---------
# =========================

WEEK_LABELS = ["So", "Mo", "Di", "Mi", "Do", "Fr", "Sa"]

def sunday_based_weekday(d: date) -> int:
    """So=0, Mo=1, ..., Sa=6 (NFC-Logik)."""
    # Python: Mon=0..Sun=6 -> So(6) soll 0 werden:
    return (d.weekday() + 1) % 7

def this_weeks_sunday(today: date) -> date:
    # most recent Sunday for NFC week
    return today - timedelta(days=(today.weekday() + 1) % 7)

def next_week_dates(today: date) -> List[date]:
    """Nächste Woche So–Sa (7 Tage) ab kommendem Sonntag."""
    last_sun = this_weeks_sunday(today)
    return [last_sun + timedelta(days=7 + i) for i in range(7)]

def normalize_name(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = " ".join(s.split())
    # Titel-Case mit Sonderzeichen ok:
    try:
        return s.title()
    except Exception:
        return s

def parse_date(x) -> Optional[date]:
    if pd.isna(x):
        return None
    # Excel-Serial oder Timestamp/String
    if isinstance(x, (pd.Timestamp, datetime)):
        return x.date()
    if isinstance(x, (float, int)):
        # Excel-Serial (pandas macht meist schon Timestamp), fallback:
        try:
            return (pd.to_datetime(x, unit="D", origin="1899-12-30")).date()
        except Exception:
            pass
    try:
        return pd.to_datetime(str(x), dayfirst=True, errors="coerce").date()
    except Exception:
        return None

def parse_time_to_minutes(x) -> Optional[int]:
    """Konvertiere Uhrzeit nach Minuten seit 00:00. Gibt None bei ungültig."""
    if pd.isna(x):
        return None
    if isinstance(x, (pd.Timestamp, datetime)):
        return x.hour * 60 + x.minute
    if isinstance(x, time):
        return x.hour * 60 + x.minute
    s = str(x).strip()
    # akzeptiere HH:MM, H:MM, HH.MM
    for sep in [":", "."]:
        if sep in s:
            parts = s.split(sep)
            if len(parts) >= 2:
                try:
                    h = int(parts[0])
                    m = int(parts[1][:2])
                    if 0 <= h < 24 and 0 <= m < 60:
                        return h * 60 + m
                except Exception:
                    return None
    # manche Excel-Zeiten kommen als Tagesbruch (0..1)
    try:
        val = float(s)
        if 0 <= val <= 1:
            mins = int(round(val * 24 * 60))
            return mins
    except Exception:
        pass
    return None

def minutes_to_hhmm(m: Optional[int]) -> str:
    if m is None:
        return ""
    h = m // 60
    mi = m % 60
    return f"{h:02d}:{mi:02d}"

def round_to_30min(m: Optional[int]) -> Optional[int]:
    if m is None:
        return None
    return int(round(m / 30) * 30)

def median_minutes(series: List[int]) -> Optional[int]:
    if not series:
        return None
    return int(np.median(series))

# =========================
# ------ Extraction -------
# =========================

def extract_records_from_excel(file) -> pd.DataFrame:
    """
    Liest ein Excel 'Touren', ab Zeile 5 (skiprows=4), ohne Header.
    Spalten (0-basiert):
    D=3, E=4, G=6, H=7, I=8 (Startzeit), O=14 (Datum)
    Gibt DataFrame mit Spalten: Name, Datum(date), StartMin(int oder None)
    """
    try:
        df = pd.read_excel(
            file,
            sheet_name="Touren",
            header=None,
            skiprows=4,   # ab Zeile 5
            dtype=object
        )
    except Exception as e:
        st.warning(f"Datei '{getattr(file, 'name', 'unbenannt')}' konnte nicht gelesen werden: {e}")
        return pd.DataFrame(columns=["Name", "Datum", "StartMin"])

    records = []
    for idx, row in df.iterrows():
        # Name-Paar 1: D+E
        n1_last = normalize_name(row[3]) if 3 in row else ""
        n1_first = normalize_name(row[4]) if 4 in row else ""
        name1 = f"{n1_last} {n1_first}".strip() if (n1_last or n1_first) else ""

        # Name-Paar 2: G+H
        n2_last = normalize_name(row[6]) if 6 in row else ""
        n2_first = normalize_name(row[7]) if 7 in row else ""
        name2 = f"{n2_last} {n2_first}".strip() if (n2_last or n2_first) else ""

        # Datum / Zeit
        dt = parse_date(row[14]) if 14 in row else None
        start_min = parse_time_to_minutes(row[8]) if 8 in row else None

        # nur Zeilen mit Datum UND mind. einem Namen
        if dt is None or (not name1 and not name2):
            continue

        # zwei Personen auf einer Tour -> zwei Datensätze
        if name1:
            records.append({"Name": name1, "Datum": dt, "StartMin": start_min})
        if name2:
            records.append({"Name": name2, "Datum": dt, "StartMin": start_min})

    if not records:
        return pd.DataFrame(columns=["Name", "Datum", "StartMin"])

    out = pd.DataFrame.from_records(records)
    # Mehrfacheinträge je Name & Datum -> früheste Zeit
    out = (
        out.groupby(["Name", "Datum"], as_index=False)
           .agg(StartMin=("StartMin", lambda s: None if all(pd.isna(s)) else np.nanmin([x for x in s if pd.notna(x)])))
    )
    return out

# =========================
# ------ Analysis ---------
# =========================

def build_typical_times(df_all: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, date]]:
    """
    df_all: Name, Datum(date), StartMin(int/None)
    Liefert:
      - DataFrame je Fahrer mit globalem Median & Median per Wochentag (So..Sa),
        Counts gesamt & je Tag
      - dict last_saturday_by_name: letzter Samstag mit Einsatz (Datum) pro Fahrer
    """
    if df_all.empty:
        cols = ["Name", "Globaler Median"] + WEEK_LABELS + ["Samples (gesamt)", "Samples je Tag (So–Sa)", "Zuletzt Samstag gearbeitet", "Bemerkung"]
        return pd.DataFrame(columns=cols), {}

    df = df_all.copy()
    df["WD"] = df["Datum"].apply(sunday_based_weekday)  # So=0..Sa=6

    results = []
    last_saturday_by_name: Dict[str, date] = {}

    for name, g in df.groupby("Name"):
        # Samples pro Tag
        per_day = {d: [int(x) for x in g.loc[(g["WD"] == d) & g["StartMin"].notna(), "StartMin"].tolist()] for d in range(7)}
        # global samples
        all_samples = [int(x) for x in g.loc[g["StartMin"].notna(), "StartMin"].tolist()]
        # Mediane
        global_med = median_minutes(all_samples)
        # Per-Tag Median (nur wenn >=3 Stichproben), sonst None
        day_meds = {}
        for d in range(7):
            day_meds[d] = median_minutes(per_day[d]) if len(per_day[d]) >= 3 else None

        # Last Saturday worked
        sat_dates = g.loc[(g["WD"] == 6) & g["StartMin"].notna(), "Datum"].tolist()
        if sat_dates:
            last_saturday_by_name[name] = max(sat_dates)

        # Counts
        total_n = len(all_samples)
        by_day_counts = [len(per_day[d]) for d in range(7)]

        row = {
            "Name": name,
            "Globaler Median": minutes_to_hhmm(global_med) if global_med is not None else "",
        }
        for d, label in enumerate(WEEK_LABELS):
            row[label] = minutes_to_hhmm(day_meds[d]) if day_meds[d] is not None else ""
        row["Samples (gesamt)"] = total_n
        row["Samples je Tag (So–Sa)"] = ", ".join([f"{lbl}:{cnt}" for lbl, cnt in zip(WEEK_LABELS, by_day_counts)])
        row["Zuletzt Samstag gearbeitet"] = last_saturday_by_name.get(name, "")
        row["Bemerkung"] = ""
        results.append(row)

    out_cols = ["Name", "Globaler Median"] + WEEK_LABELS + ["Samples (gesamt)", "Samples je Tag (So–Sa)", "Zuletzt Samstag gearbeitet", "Bemerkung"]
    out_df = pd.DataFrame(results, columns=out_cols).sort_values("Name").reset_index(drop=True)
    return out_df, last_saturday_by_name

def build_recommendations(typ_df: pd.DataFrame, last_sat_map: Dict[str, date], today: date) -> pd.DataFrame:
    """Erzeuge Einsatzplan Folgewoche (So–Sa) mit HH:MM oder leer."""
    week_dates = next_week_dates(today)  # So..Sa
    date_headers = [f"{lbl} {d.strftime('%d.%m.')}" for lbl, d in zip(WEEK_LABELS, week_dates)]

    rows = []
    # letzte & vorletzte Samstage:
    last_sunday = this_weeks_sunday(today)
    last_saturday = last_sunday + timedelta(days=6)  # letzter Samstag (in der aktuellen Woche)
    prev_saturday = last_saturday - timedelta(days=7)

    for _, r in typ_df.iterrows():
        name = r["Name"]
        # globaler Median in Minuten
        global_med = None
        if r["Globaler Median"]:
            hh, mm = map(int, r["Globaler Median"].split(":"))
            global_med = hh * 60 + mm

        # Per-Tag Minuten
        per_day_min = {}
        for idx, lbl in enumerate(WEEK_LABELS):
            v = r[lbl]
            per_day_min[idx] = None
            if isinstance(v, str) and v:
                try:
                    h, m = map(int, v.split(":"))
                    per_day_min[idx] = h * 60 + m
                except Exception:
                    pass

        last_sat = last_sat_map.get(name)
        # Bemerkung aus Rotation?
        rotation_hit = last_sat == last_saturday  # hat zuletzt (gerade) am Sa gearbeitet

        row = {"Name": name}
        for d_idx, (lbl, dt) in enumerate(zip(WEEK_LABELS, week_dates)):
            # Regel: Per-Tag Median (wenn vorhanden), sonst globaler Median (wenn vorhanden), sonst leer
            cand = per_day_min.get(d_idx)
            if cand is None:
                cand = global_med
            cand = round_to_30min(cand)

            # Rotation: wenn kommender Tag Sa und zuletzt Sa gearbeitet -> leer lassen
            if d_idx == 6 and rotation_hit:
                row[f"{lbl} {dt.strftime('%d.%m.')}"] = ""
            else:
                row[f"{lbl} {dt.strftime('%d.%m.')}"] = minutes_to_hhmm(cand)

        rows.append(row)

    rec_df = pd.DataFrame(rows, columns=["Name"] + date_headers).sort_values("Name").reset_index(drop=True)

    # Bemerkung im typ_df setzen (für Transparenz)
    mask = typ_df["Name"].isin([n for n, d in last_sat_map.items() if d == last_saturday])
    typ_df.loc[mask, "Bemerkung"] = typ_df.loc[mask, "Bemerkung"].astype(str).str.replace(
        r"$", "Samstagsrotation: zuletzt gearbeitet am letzten Samstag → kommenden Sa freihalten.", regex=True
    )
    return rec_df

def format_excel_two_sheets(rec_df: pd.DataFrame, typ_df: pd.DataFrame, today: date) -> bytes:
    """Erzeuge Excel (Bytes) mit Formatierung und zwei Blättern."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Sheet 1
    ws1 = wb.active
    ws1.title = "Einsatz Folgewoche"
    # Headers
    for j, col in enumerate(rec_df.columns, start=1):
        ws1.cell(row=1, column=j, value=col)

    # Data
    for i, (_, row) in enumerate(rec_df.iterrows(), start=2):
        for j, col in enumerate(rec_df.columns, start=1):
            ws1.cell(row=i, column=j, value=row[col])

    # Sheet 2
    ws2 = wb.create_sheet("Typische Startzeiten")
    for j, col in enumerate(typ_df.columns, start=1):
        ws2.cell(row=1, column=j, value=col)
    for i, (_, row) in enumerate(typ_df.iterrows(), start=2):
        for j, col in enumerate(typ_df.columns, start=1):
            ws2.cell(row=i, column=j, value=row[col])

    # Styling
    header_fill = PatternFill("solid", fgColor="EEECE1")
    saturday_fill = PatternFill("solid", fgColor="F8D7DA")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin", color="BBBBBB"),
        right=Side(style="thin", color="BBBBBB"),
        top=Side(style="thin", color="BBBBBB"),
        bottom=Side(style="thin", color="BBBBBB"),
    )

    # Sheet 1 widths + header style
    widths1 = [26] + [14] * (len(rec_df.columns) - 1)
    for col_idx, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = w
        c = ws1.cell(row=1, column=col_idx)
        c.font = header_font; c.alignment = center; c.fill = header_fill; c.border = border

    # Saturday highlight (last column)
    if len(rec_df.columns) >= 2:
        saturday_col_ws1 = len(rec_df.columns)
        max_rows = max(2, rec_df.shape[0] + 20)
        for r in range(1, max_rows + 1):
            c = ws1.cell(row=r, column=saturday_col_ws1)
            c.fill = saturday_fill
            c.border = border
            if r == 1:
                c.font = header_font
                c.alignment = center

    # Data borders + align
    for r in range(2, rec_df.shape[0] + 2):
        name_cell = ws1.cell(row=r, column=1); name_cell.alignment = left; name_cell.border = border
        for cidx in range(2, len(rec_df.columns) + 1):
            cell = ws1.cell(row=r, column=cidx)
            cell.alignment = center
            if cidx != len(rec_df.columns):
                cell.border = border

    # Sheet 2 widths + header style
    widths2 = [26, 16] + [12]*7 + [18, 28, 24, 36]
    for col_idx, w in enumerate(widths2[:len(typ_df.columns)], start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font; c.alignment = center; c.fill = header_fill; c.border = border

    # Saturday highlight in sheet 2 (Name, Globaler Median, So..Sa => Sa an Position 10)
    if len(typ_df.columns) >= 10:
        saturday_col_ws2 = 10
        max_rows2 = max(2, typ_df.shape[0] + 20)
        for r in range(1, max_rows2 + 1):
            c = ws2.cell(row=r, column=saturday_col_ws2)
            c.fill = saturday_fill
            c.border = border
            if r == 1:
                c.font = header_font
                c.alignment = center

    # Align data
    for r in range(2, typ_df.shape[0] + 2):
        ws2.cell(row=r, column=1).alignment = left  # Name
        for cidx in range(2, len(typ_df.columns) + 1):
            ws2.cell(row=r, column=cidx).alignment = center
            ws2.cell(row=r, column=cidx).border = border

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

# =========================
# -------- UI -------------
# =========================

st.set_page_config(page_title="Dienstplan → Folgewoche (Excel)", layout="wide")
st.title("🧭 Dienstplan-Auswertung → Einsatz-Empfehlung Folgewoche (So–Sa)")

st.markdown("""
Lade **eine oder mehrere** Excel-Dateien hoch. Erwartetes Schema je Datei:
- Blatt **„Touren“**
- **ab Zeile 5** (Zeilen 1–4 Kopf/Meta)
- **D+E** _oder_ **G+H** = Nachname + Vorname (beide Paare werden geprüft; je Paar ein Datensatz)
- **O** = Datum · **I** = Startzeit
""")

files = st.file_uploader("Excel-Dateien auswählen", type=["xlsx", "xls"], accept_multiple_files=True)

col_a, col_b = st.columns([1, 1])
with col_a:
    today_override = st.date_input("Stichtag (heute)", value=date.today())
with col_b:
    min_samples_weekday = st.slider("Min. Stichproben für Tages-Median", 1, 5, 3, help="Unterhalb wird der globale Median genutzt, wenn vorhanden.")
rotation_enabled = st.checkbox("Samstags-Rotation (wer letzten Samstag arbeitete, bleibt kommenden Sa leer)", value=True)

st.markdown("---")

if st.button("Analysieren & Excel erzeugen", type="primary"):
    if not files:
        st.warning("Bitte mindestens eine Excel-Datei auswählen.")
        st.stop()

    # 1) Extraction über alle Dateien
    all_df_list = []
    for f in files:
        part = extract_records_from_excel(f)
        if not part.empty:
            all_df_list.append(part)

    if not all_df_list:
        st.error("Keine verwertbaren Datensätze gefunden (prüfe Blattname/Spalten/Zeilen).")
        st.stop()

    df_all = pd.concat(all_df_list, ignore_index=True)

    # 2) Typische Startzeiten + letzter Samstag je Fahrer
    typ_df_raw, last_sat_map = build_typical_times(df_all)

    # ggf. Schwellwert für Tages-Median anpassen (Standard=3)
    # (build_typical_times nutzt fix >=3; hier setzen wir optional dynamisch um)
    # Für Einfachheit: wir rechnen die Anzeige nicht neu, aber die Empfehlungen können
    # den Schwellwert berücksichtigen, indem wir r["So"].. überschreiben, wenn counts < slider.
    # -> Dazu Counts aus "Samples je Tag" parsen und leeren Tagesmedian, wenn zu klein:

    def apply_min_samples(typ_df: pd.DataFrame, min_samples: int) -> pd.DataFrame:
        df = typ_df.copy()
        # parse counts per day
        for idx, row in df.iterrows():
            counts_map = {p.split(":")[0]: int(p.split(":")[1]) for p in str(row["Samples je Tag (So–Sa)"]).split(", ")}
            for lbl in WEEK_LABELS:
                if counts_map.get(lbl, 0) < min_samples:
                    df.at[idx, lbl] = ""  # Tagesmedian verwerfen
        return df

    typ_df = apply_min_samples(typ_df_raw, min_samples_weekday)

    # 3) Empfehlungen bauen
    rec_df = build_recommendations(typ_df.copy(), last_sat_map if rotation_enabled else {}, today_override)

    # 4) Download-Datei erzeugen
    excel_bytes = format_excel_two_sheets(rec_df, typ_df, today_override)
    filename = f"Einsatzplan_Folgewoche_{today_override.strftime('%Y-%m-%d')}.xlsx"

    st.success("Fertig! Du kannst die Excel-Datei jetzt herunterladen.")
    st.download_button("💾 Excel herunterladen", data=excel_bytes, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Vorschau
    st.subheader("Vorschau: Einsatz Folgewoche")
    st.dataframe(rec_df, use_container_width=True)
    st.subheader("Vorschau: Typische Startzeiten")
    st.dataframe(typ_df, use_container_width=True)
